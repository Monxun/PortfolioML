from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging
import os
import string
import streamlit as st
import pandas as pd
from services import (
    files_dir,
    files,  
    logging, 
    months, 
    years, 
    check_file,
    get_current, 
    get_summary, 
    get_voc,
    log_summary, 
    log_voc,
    show_summary, 
    show_voc,
    show_logs
)

#####################################################
#NAVIGATION

st.image('./stocks.png', width=100)
st.title('Portfolio Analysis ML ')
pages = ['Select by File', 'Log File']
navigation = st.selectbox('Select a page', pages)

st.write('_' * 30)

st.header(navigation)

data_flag = False

if navigation != 'Log File':

    if navigation == 'Select by File':

        #####################################################
        #SELECT FILE

        # workbook file selectbox 
        file_selector = st.selectbox('Select a file', files)  # replace with streamlit selector 
        file_name =' '.join([item for item in file_selector.replace('.', '_').split("_") if item != 'csv' or 'xlsx'])

        # verify file exists / error handling
        check_file(file_selector)

        # if selected file is a .xlsx then display text and load openpyxl
        if '.xlsx' in file_selector:
            st.text(f'({file_selector} is selected.)')
            wb = load_workbook(f'data/{file_selector}')
            logging.info(f'({file_selector} is selected.)')

        if '.csv' in file_selector:
            st.text(f'({file_selector} is selected.)')
            data = pd.read_csv(f'data/{file_selector}')

            if file_selector == 'Webull_Orders_Records_Options.csv':
                df = data[['Symbol', 'Side', 'Status', 'Filled', 'Price', 'Avg Price']]
            else:
                df = data

            logging.info(f'({file_selector} is selected.)')

        st.write('_' * 30)
        st.subheader(f'Data for {file_name}')

        if 'Options' in file_selector:
            # analysis selector
            analysis_type = ['Orders', 'Describe', 'Data', 'Placed / Filled']
            analysis_selector = st.selectbox('Select Analysis', analysis_type)

            filled_filter = df['Status'] == 'Filled'
            df['Symbol'] = df['Symbol'].astype(str).str[:6]
            df['Symbol'] = df['Symbol'].astype(str).str.rstrip('0123456789')

            if analysis_selector == 'Orders':
                st.dataframe(df[filled_filter])

            if analysis_selector == 'Describe':
                st.write(df.describe())

            if analysis_selector == 'Data':
                st.dataframe(data)

            if analysis_selector == 'Placed / Filled':
                pf = data[filled_filter][['Name', 'Placed Time', 'Filled Time']]
                pf['Name'] = pf['Name'].astype(str).str[0:17]
                pf['Name'] = pf['Name'].astype(str).str.rstrip('0123456789')
                pf['Placed Time'] = pd.to_datetime(pf['Placed Time'], infer_datetime_format=True).astype(str).str[11:20]
                pf['Filled Time'] = pd.to_datetime(pf['Filled Time'], infer_datetime_format=True).astype(str).str[11:20]
                pf['Date'] = pf['Name'].astype(str).str[-11:]
                pf['Name'] = pf['Name'].astype(str).str[0:-11]
                pf['Name'] = pf['Name'].astype(str).str.replace('[^a-zA-Z]', '')

                cols = ['Name', 'Date', 'Placed Time', 'Filled Time']
                pf = pf[cols]
       
                st.dataframe(pf)

                st.subheader('Stocks Picked This Year')
                stocks = list(set(pf['Name']))
                st.write(stocks)
                

                



else:

    #####################################################
    #LOG FILE

    show_logs()




